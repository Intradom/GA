"""

Symbolic regression using instruction set approach

Valid operators: +, -, *, / 

Valid operands (suppose x is an input variable): const, x, cos(x), sin(x)

"""

import sys
import numpy as np
import openpyxl as op

POSSIBLE_TYPES = ["load", "cos", "sin", "add", "sub", "mul", "div"]
BI_OP_START = 3

CONST_MAX = 10 # Loaded constants will range from -CONST_MAX to CONST_MAX
INITIAL_INSTRUCTIONS = 10

class Instruction(object):
    def __init__(self, i_num, var):
        if var:
            self.itype = "loadVar" # Will never mutate these instructions
            self.op1 = i_num # Later dictates what variable num it indexed into variables array
            self.op2 = None
        else:
            self.mutate(i_num)

    def mutate(self, i_num):
        r = 0
    
        if i_num == 0:
            self.itype = "load"
        else:
            r = np.random.randint(len(POSSIBLE_TYPES))
            self.itype = POSSIBLE_TYPES[r]
        
        if self.itype == "load":
            self.op1 = np.float16((np.random.rand() - 0.5) * 2 * CONST_MAX)
            self.op2 = None
        else:              
            if r < BI_OP_START: # Requires only 1 operand
                self.op1 = np.random.randint(i_num)
                self.op2 = None
            else: # Requires two operands
                self.op1 = np.random.randint(i_num)
                self.op2 = np.random.randint(i_num)
    
    def getType(self):
        return self.itype 
    
    def getOps(self):
        return self.op1, self.op2
        
    def print_instruction(self, num):
        if self.var:
            print(str(num) + ". " + self.itype + ": " + "X_" + str(self.op1) + ", " + str(self.op2))
        else:
            print(str(num) + ". " + self.itype + ": " + str(self.op1) + ", " + str(self.op2))

# For debugging
def print_instuctions(num_vars, instructions):
    for i in range(instructions.size):
        instructions[i].print_instruction(i)
        
def eval_instructions(itype, val1, val2 = None):
    if itype == "load":
        return val1
    elif itype == "cos":
        return math.cos(val1)
    elif itype == "sin":
        return math.sin(val1)
    elif itype == "add":
        return val1 + val2
    elif itype == "sub":
        return val1 - val2
    elif itype == "mul":
        return val1 * val2
    elif itype == "div":
        return val1 / float(val2) # Just in case variable read returned an integer
    else:
        print("Unidentified instruction")
        sys.exit()
    
        
def eval_fitness(sheet, instructions):
    rmse = 0.0
    results = np.empty([instructions.size, 1]) # No need to reset because values will always freshly set on each row in a descending manner
    for r in range(sheet.max_row - 2):
        i = r + 3 # Remove the headers
        y_true = sheet.cell(row = i, column = 1).value
        
        # Get predicted y value
        for i in range(instructions.size):
            op1, op2 = instructions[i].getOps()
            if instructions[i].getType() == "loadVar":
                results[i] = sheet.cell(row = i, column = 2 + op1)
            else:
                if op2 is None:
                    results[i] = eval_instructions(instructions[i].getType(), results[op1])
                else:
                    results[i] = eval_instructions(instructions[i].getType(), results[op1], results[op2])
                    
        rmse += math.pow(y_true - results[results.size - 1], 2)
        
    rmse /= float(sheet.max_row - 2)
    return math.sqrt(rmse)

def main():
    # Initialize
    wb = op.load_workbook("gen_data_for_testing.xlsx")
    sheet = wb.get_sheet_by_name('Exp1')
    num_vars = sheet.max_column - 1
    # Copying over x and y saves time from constant reads of excel file? TODO: Check on this
    
    instructions = np.empty([0, 1])

    for i in range(num_vars):
        instructions = np.append(instructions, Instruction(i, True))
    
    for i in range(INITIAL_INSTRUCTIONS):
        instructions = np.append(instructions, Instruction(i + num_vars, False))
       
    # DEBUG 
    print_instuctions(num_vars, instructions)
        
    # Evaluate fitness (RMSE)
    print eval_fitness(sheet, instructions)

    """ Loop """
    # Mutate

    # Evaluate fitness (RMSE)

    # If pass stopping criteria, exit loop. Else (end regardless of result if time exceeded), revert if mutation made worse and loop; otherwise, keep changes and loop again

if __name__ == "__main__":
    main()
